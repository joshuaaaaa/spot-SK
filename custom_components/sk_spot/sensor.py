"""SK Spot sensor."""
from datetime import datetime, timedelta, time
from random import randint
import logging
import io

from openpyxl import load_workbook
import aiohttp

from homeassistant.components.sensor import SensorEntity, SensorStateClass
from homeassistant.config_entries import ConfigEntry
from homeassistant.core import HomeAssistant
from homeassistant.helpers import event
from homeassistant.helpers.entity_platform import AddEntitiesCallback
from homeassistant.helpers.update_coordinator import (
    CoordinatorEntity,
    DataUpdateCoordinator,
    UpdateFailed,
)
from homeassistant.util import dt as dt_util

from .const import DOMAIN, CONF_UNIT, UNIT_MWH, UNIT_KWH

_LOGGER = logging.getLogger(__name__)

# Čas, kdy by data měla být publikována (13:05 slovenského času)
DATA_AVAILABLE_TIME = time(13, 5)
# Náhodné zpoždění (0-120 sekund) pro prevenci synchronizace všech uživatelů
JITTER_SECONDS = 120


async def async_setup_entry(
    hass: HomeAssistant,
    entry: ConfigEntry,
    async_add_entities: AddEntitiesCallback,
) -> None:
    """Setup senzoru."""
    coordinator = SKSpotCoordinator(hass)
    await coordinator.async_config_entry_first_refresh()
    async_add_entities([SKSpotSensor(coordinator, entry)])
    # Naplánuj první automatickou aktualizaci
    coordinator.schedule_next_update()


class SKSpotCoordinator(DataUpdateCoordinator):
    """Coordinator pro stahování dat."""

    def __init__(self, hass: HomeAssistant) -> None:
        """Init."""
        super().__init__(
            hass,
            _LOGGER,
            name="SK Spot",
        )
        self._update_schedule = None  # Handle pro naplánovanou aktualizaci
        self._last_download_date = None
        self._today_prices = {}
        self._tomorrow_prices = {}
        self._tomorrow_available = False

    def _validate_price_data(self, prices):
        """Zkontroluj zda jsou data validní (máme všech 96 záznamů pro celý den)."""
        if not prices:
            return False
        # Měli bychom mít 96 záznamů (24 hodin * 4 čtvrthodiny)
        # Kontrola, že máme alespoň 90 záznamů (95% úplnosti)
        if len(prices) < 90:
            return False
        return True

    def has_tomorrow_data(self) -> bool:
        """Zkontroluj, zda máme data pro zítřek."""
        if not self._tomorrow_prices:
            return False
        # Ověř, že máme validní data (alespoň 90 záznamů)
        return self._validate_price_data(self._tomorrow_prices)

    def schedule_next_update(self):
        """Naplánuj další aktualizaci dat."""
        from zoneinfo import ZoneInfo

        # Zruš předchozí naplánovanou aktualizaci, pokud existuje
        if self._update_schedule is not None:
            self._update_schedule()
            self._update_schedule = None

        utc = ZoneInfo("UTC")
        bratislava_tz = ZoneInfo("Europe/Bratislava")
        now_bratislava = dt_util.now(bratislava_tz)

        if self.has_tomorrow_data():
            # Už máme data pro zítřek, další update bude zítra po 13:05
            local_target = datetime.combine(
                (now_bratislava + timedelta(days=1)).date(),
                DATA_AVAILABLE_TIME,
                tzinfo=bratislava_tz,
            )
            local_target += timedelta(seconds=randint(1, JITTER_SECONDS))
            _LOGGER.info("Máme zítřejší data, další update: %s", local_target)
        else:
            # Nemáme data pro zítřek
            if DATA_AVAILABLE_TIME < now_bratislava.time():
                # Čas 13:05 už uplynul dnes, ale nemáme zítřejší data
                # Zkusíme to znovu za 5 minut
                local_target = now_bratislava + timedelta(minutes=5)
                _LOGGER.info("Nemáme zítřejší data, další pokus za 5 minut: %s", local_target)
            else:
                # Ještě není 13:05 dnes, naplánuj update na 13:05
                local_target = datetime.combine(
                    now_bratislava.date(),
                    DATA_AVAILABLE_TIME,
                    tzinfo=bratislava_tz,
                )
                local_target += timedelta(seconds=randint(1, JITTER_SECONDS))
                _LOGGER.info("Další update dnes ve: %s", local_target)

        # Převeď na UTC (správně ošetří letní čas)
        utc_time = local_target.astimezone(utc)

        # Naplánuj aktualizaci
        self._update_schedule = event.async_track_point_in_utc_time(
            hass=self.hass,
            action=self._on_schedule,
            point_in_time=utc_time,
        )

        return utc_time

    async def _on_schedule(self, _):
        """Callback pro naplánovanou aktualizaci."""
        _LOGGER.info("Spouštím naplánovanou aktualizaci")
        await self.async_request_refresh()
        # Naplánuj další update
        self.schedule_next_update()

    async def _async_update_data(self):
        """Stáhni data."""
        now = dt_util.now()
        today = now.date()

        # Pokud se změnil den (po půlnoci), posuň zítřejší ceny na dnešní
        if self._last_download_date is not None and self._last_download_date < today:
            _LOGGER.info("Den se změnil z %s na %s - posouváme zítřejší ceny na dnešní",
                        self._last_download_date, today)
            # Přesuň včerejší "zítřejší" ceny na dnešní "dnešní" ceny
            if self.has_tomorrow_data():
                self._today_prices = self._tomorrow_prices.copy()
                _LOGGER.info("Přesunuto %d zítřejších cen na dnešní", len(self._today_prices))
            else:
                # Pokud jsme neměli zítřejší ceny, vynuluj dnešní
                self._today_prices = {}
            # Vyčisti zítřejší data
            self._tomorrow_prices = {}
            self._tomorrow_available = False
            # Nastav, že jsme ještě dnes nestahovali
            self._last_download_date = None

        # Pokud ještě dnes nestahovali, stáhni data
        should_download = (self._last_download_date != today)

        if should_download:
            try:
                await self._fetch_prices(today)
                self._last_download_date = today
                _LOGGER.info("Staženo nových cen pro dnes (%d) a zítra (%s)",
                           len(self._today_prices),
                           f"{len(self._tomorrow_prices)} - dostupné" if self.has_tomorrow_data() else "nedostupné")
            except Exception as err:
                _LOGGER.error("Chyba při stahování: %s", err)
                # Pokud nemáme vůbec žádná data, vyvolej chybu
                if not self._validate_price_data(self._today_prices):
                    raise UpdateFailed(f"Chyba: {err}") from err
                # Jinak pokračuj se starými daty
                _LOGGER.warning("Používám stará data")

        # Určení aktuální ceny podle 15minutového intervalu
        current_hour = now.hour
        current_minute = now.minute
        # Vypočítat index 15minutového intervalu (0-95)
        quarter_index = (current_hour * 4) + (current_minute // 15)

        # Pokud je dnes a máme data
        if self._today_prices:
            current_price = self._today_prices.get(quarter_index)
        else:
            current_price = 0

        return {
            "current_price": current_price if current_price is not None else 0,
            "today_prices": self._today_prices,
            "tomorrow_prices": self._tomorrow_prices if self.has_tomorrow_data() else {},
            "tomorrow_available": self.has_tomorrow_data(),
            "last_update": now.isoformat(),
        }

    async def _fetch_prices(self, today):
        """Stáhni a parsuj XLSX pro dnes a zítra."""
        tomorrow = today + timedelta(days=1)

        # Stáhnout dnešní ceny
        try:
            self._today_prices = await self._fetch_day_prices(today)
            _LOGGER.info("Dnešní ceny úspěšně staženy: %d záznamů", len(self._today_prices))
        except Exception as err:
            _LOGGER.warning("Nelze stáhnout dnešní ceny: %s", err)
            self._today_prices = {}
            raise  # Propaguj chybu pokud ani dnes nejde stáhnout

        # Stáhnout zítřejší ceny
        try:
            self._tomorrow_prices = await self._fetch_day_prices(tomorrow)
            self._tomorrow_available = True
            _LOGGER.info("Zítřejší ceny úspěšně staženy: %d záznamů", len(self._tomorrow_prices))
        except Exception as err:
            _LOGGER.info("Zítřejší ceny ještě nejsou dostupné: %s", err)
            self._tomorrow_prices = {}
            self._tomorrow_available = False

    async def _fetch_day_prices(self, date):
        """Stáhni ceny pro konkrétní den."""
        delivery_date = date.strftime("%Y-%m-%d")

        url = (
            f"https://isot.okte.sk/api/v1/dam/report/detailed"
            f"?lang=sk-SK"
            f"&deliverydayfrom={delivery_date}"
            f"&deliverydayto={delivery_date}"
            f"&format=xlsx"
        )

        _LOGGER.debug("Stahuji data pro %s z: %s", delivery_date, url)

        timeout = aiohttp.ClientTimeout(total=60)
        async with aiohttp.ClientSession(timeout=timeout) as session:
            async with session.get(url) as response:
                if response.status != 200:
                    _LOGGER.error("API vrátilo HTTP %d pro %s", response.status, delivery_date)
                    raise UpdateFailed(f"HTTP {response.status}")
                content = await response.read()
                _LOGGER.debug("Staženo %d bytů pro %s", len(content), delivery_date)

        # Parse XLSX
        workbook = load_workbook(filename=io.BytesIO(content), data_only=True)
        sheet = workbook.active

        prices = {}

        # Sloupec K = 11. sloupec (index 11)
        # Data začínají od řádku 2
        # Máme 96 řádků (24h * 4 čtvrthodiny)
        for row_idx, row in enumerate(sheet.iter_rows(min_row=2, min_col=11, max_col=11), start=0):
            cell = row[0]
            if cell.value is not None and row_idx < 96:
                try:
                    price = float(cell.value)
                    prices[row_idx] = round(price, 4)
                except (ValueError, TypeError):
                    _LOGGER.warning("Nelze parsovat cenu na řádku %d pro %s: %s",
                                   row_idx + 2, delivery_date, cell.value)
                    continue

        if not prices:
            _LOGGER.error("XLSX pro %s neobsahuje žádná data", delivery_date)
            raise UpdateFailed("Žádná data v XLSX")

        _LOGGER.debug("Naparsováno %d cen pro %s", len(prices), delivery_date)
        return prices


class SKSpotSensor(CoordinatorEntity, SensorEntity):
    """SK Spot Price sensor."""

    _attr_name = "SK Spot Price"
    _attr_state_class = SensorStateClass.MEASUREMENT

    def __init__(self, coordinator: SKSpotCoordinator, entry: ConfigEntry) -> None:
        """Init."""
        super().__init__(coordinator)
        self._attr_unique_id = f"{entry.entry_id}_price"
        self._entry = entry
        self._unit = entry.data.get(CONF_UNIT, UNIT_MWH)

    @property
    def native_unit_of_measurement(self):
        """Jednotka měření."""
        if self._unit == UNIT_KWH:
            return "EUR/kWh"
        return "EUR/MWh"

    @property
    def native_value(self):
        """Aktuální cena."""
        if self.coordinator.data is None:
            return 0
        
        price = self.coordinator.data.get("current_price", 0)
        
        # Převod MWh -> kWh (dělit 1000)
        if self._unit == UNIT_KWH:
            return round(price / 1000, 6)
        
        return price

    @property
    def extra_state_attributes(self):
        """Atributy."""
        if self.coordinator.data is None:
            return {}
        
        today_prices = self.coordinator.data.get("today_prices", {})
        tomorrow_prices = self.coordinator.data.get("tomorrow_prices", {})
        tomorrow_available = self.coordinator.data.get("tomorrow_available", False)
        
        now = dt_util.now()
        today_date = now.date()
        tomorrow_date = today_date + timedelta(days=1)
        
        # Vytvoření jednoho slovníku pro dnes i zítra
        all_prices = {}
        
        # Přidat dnešní ceny
        for idx in range(96):
            hour = idx // 4
            minute = (idx % 4) * 15
            
            if idx in today_prices:
                dt = datetime.combine(today_date, time(hour, minute))
                dt = dt_util.as_local(dt_util.as_utc(dt))
                iso_time = dt.isoformat()
                
                price = today_prices[idx]
                if self._unit == UNIT_KWH:
                    price = round(price / 1000, 2)
                else:
                    price = round(price, 2)
                    
                all_prices[iso_time] = price
        
        # Přidat zítřejší ceny - POUZE pokud jsou skutečně dostupné A je po 13:00
        # (Data se zveřejňují každý den ve 13:00)
        if tomorrow_available and tomorrow_prices and now.time() >= time(13, 0):
            for idx in range(96):
                hour = idx // 4
                minute = (idx % 4) * 15

                if idx in tomorrow_prices:
                    price = tomorrow_prices[idx]

                    dt = datetime.combine(tomorrow_date, time(hour, minute))
                    dt = dt_util.as_local(dt_util.as_utc(dt))
                    iso_time = dt.isoformat()

                    if self._unit == UNIT_KWH:
                        price = round(price / 1000, 2)
                    else:
                        price = round(price, 2)

                    all_prices[iso_time] = price

        _LOGGER.debug("Atributy obsahují %d záznamů (dnes: %d, zítra: %d, zítra dostupné: %s)",
                     len(all_prices), len(today_prices), len(tomorrow_prices), tomorrow_available)
        return all_prices
